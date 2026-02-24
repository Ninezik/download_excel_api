from fastapi import FastAPI, Query, HTTPException
from fastapi.responses import Response
import pandas as pd
import psycopg2
import pymssql
import requests
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from io import BytesIO
from datetime import datetime, timedelta
from PIL import Image as PILImage
from concurrent.futures import ThreadPoolExecutor, as_completed

app = FastAPI()

# =========================
# KONFIG REDSHIFT
# =========================
DB_HOST = "pos-redshift.cwig526q7i0q.ap-southeast-3.redshift.amazonaws.com"
DB_PORT = "5439"
DB_NAME = "posind_kurlog"
DB_USER = "rda_analis"
DB_PASSWORD = "GcTz69eZ6UwNnRhypjx9Ysk8"


# =========================
# KONFIG sqlserver
# =========================
DB_HOST_2 = "bansosreport-db.cmfru4yoszrg.ap-southeast-3.rds.amazonaws.com"
DB_PORT_2 = "1433"
DB_NAME_2 = "DB_REFERENSI"
DB_USER_2 = "admin"
DB_PASSWORD_2 = "B4ns05dB"

CHUNK_SIZE = 100  # maksimal 100 row per sheet


@app.get("/download")
def download_excel(
    customer_code: str = Query(..., description="Customer Code"),
    start_date: str = Query(..., description="Format: YYYYMMDD")
):
    conn_2 = pymssql.connect(
    server=DB_HOST_2,
    user=DB_USER_2,
    password=DB_PASSWORD_2,
    database=DB_NAME_2,
    port=int(DB_PORT_2)
    )
    cursor_2 = conn_2.cursor()
    cursor_2.execute("SELECT TOP 5 * FROM daftar_customer_code_download")
    data_set = {row[0] for row in cursor_2.fetchall()}

    # =========================
    # VALIDASI CUSTOMER
    # =========================
    if customer_code not in  data_set:
        raise HTTPException(
            status_code=403,
            detail="Pelanggan harus didaftarkan terlebih dahulu, hubungi admin."
        )

    # =========================
    # VALIDASI TANGGAL
    # =========================
    try:
        start_dt = datetime.strptime(start_date, "%Y%m%d")
        end_dt = start_dt + timedelta(days=1)
        start_date_sql = start_dt.strftime("%Y-%m-%d")
        end_date_sql = end_dt.strftime("%Y-%m-%d")
    except ValueError:
        raise HTTPException(status_code=400, detail="Format tanggal harus YYYYMMDD")

    # =========================
    # CONNECT REDSHIFT
    # =========================
    conn = psycopg2.connect(
        host=DB_HOST,
        port=DB_PORT,
        dbname=DB_NAME,
        user=DB_USER,
        password=DB_PASSWORD
    )

    query = """
    SELECT 
        t1.connote__connote_code,
        t1.customer_code,
        t1.connote__connote_receiver_name,
        t1.connote__connote_receiver_address_detail,
        t1.connote__connote_state,
        t1.pod__timereceive,
        t2.pod__photo,
        t2.pod__signature
    FROM nipos.nipos t1
    JOIN nipos.nipos_pod_url t2
        ON t1.connote__connote_code = t2.connote__connote_code
    WHERE t1.customer_code = %s
        AND t1.connote__created_at >= %s
        AND t1.connote__created_at < %s
        AND t1.connote__connote_state IN ('DELIVERED (RETURN DELIVERY)','DELIVERED')
    """

    df = pd.read_sql(query, conn, params=(customer_code, start_date_sql, end_date_sql))
    conn.close()

    if df.empty:
        return {"message": "Data tidak ditemukan"}

    # Batasi maksimal data (proteksi)
    if len(df) > 2500:
        raise HTTPException(
            status_code=400,
            detail="Maksimal 2500 data per request."
        )

    # =========================
    # BUAT WORKBOOK
    # =========================
    wb = Workbook()
    wb.remove(wb.active)

    chunks = [df[i:i + CHUNK_SIZE] for i in range(0, len(df), CHUNK_SIZE)]
    session = requests.Session()

    for sheet_index, chunk in enumerate(chunks, start=1):

        ws = wb.create_sheet(title=f"Data_{sheet_index}")
        ws.append(chunk.columns.tolist())

        # Tulis data tanpa gambar dulu
        for _, row in chunk.iterrows():
            row_data = []
            for col in chunk.columns:
                if col in ["pod__photo", "pod__signature"]:
                    row_data.append("")
                else:
                    row_data.append(row[col])
            ws.append(row_data)

        def insert_image_from_url(url, cell):
            if pd.isna(url):
                return
            try:
                response = session.get(
                    url,
                    headers={"User-Agent": "Mozilla/5.0"},
                    timeout=3
                )

                if response.status_code == 200:

                    img = PILImage.open(BytesIO(response.content))

                    if img.mode in ("RGBA", "P"):
                        img = img.convert("RGB")

                    # Resize lebih kecil
                    img.thumbnail((250, 250))

                    compressed = BytesIO()
                    img.save(
                        compressed,
                        format="JPEG",
                        quality=30,
                        optimize=True,
                        progressive=True
                    )
                    compressed.seek(0)

                    excel_img = Image(compressed)

                    max_display = 90
                    ratio = min(
                        max_display / excel_img.width,
                        max_display / excel_img.height
                    )
                    excel_img.width = int(excel_img.width * ratio)
                    excel_img.height = int(excel_img.height * ratio)

                    ws.add_image(excel_img, cell)

            except Exception:
                pass

        photo_col = chunk.columns.get_loc("pod__photo") + 1
        sign_col = chunk.columns.get_loc("pod__signature") + 1

        # =========================
        # PARALLEL INSERT IMAGE
        # =========================
        def process_row(i):
            excel_row = i + 2
            ws.row_dimensions[excel_row].height = 85

            photo_url = chunk.iloc[i]["pod__photo"]
            sign_url = chunk.iloc[i]["pod__signature"]

            photo_cell = ws.cell(row=excel_row, column=photo_col).coordinate
            sign_cell = ws.cell(row=excel_row, column=sign_col).coordinate

            insert_image_from_url(photo_url, photo_cell)
            insert_image_from_url(sign_url, sign_cell)

        with ThreadPoolExecutor(max_workers=15) as executor:
            futures = [executor.submit(process_row, i) for i in range(len(chunk))]
            for _ in as_completed(futures):
                pass

        ws.column_dimensions[
            ws.cell(row=1, column=photo_col).column_letter
        ].width = 20

        ws.column_dimensions[
            ws.cell(row=1, column=sign_col).column_letter
        ].width = 20

    # =========================
    # RETURN FILE
    # =========================
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"data_{customer_code}_{start_date}.xlsx"

    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )
